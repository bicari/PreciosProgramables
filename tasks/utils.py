import pandas as pd
from .dbisam import DBISAMDatabase
import win32print
from .labels import PrinterLabel
from django.conf import settings
import os
import logging
from .models import ProductsTasks, ListaEtiquetaDetalle
from datetime import datetime, time
import time as t
import openpyxl
from openpyxl.styles import Font, PatternFill
import socket
from uuid import uuid4

logger = logging.getLogger(__name__)

TCP_IP = '10.150.1.122'  # IP de la impresora
TCP_PORT = 9100


def detectar_skus_repetidos(df: pd.DataFrame, header: int) -> dict:
    """Detecta SKU que aparecen más de una vez en el DataFrame.

    Args:
        df: DataFrame del Excel ya parseado (con dropna aplicado).
        header: número de fila del encabezado (1-based) para calcular la línea real del Excel.

    Returns:
        Dict {sku: [lineas_excel]} sólo para los SKU con más de una aparición.
    """
    sku_lineas: dict = {}
    for fila in df.itertuples():
        sku = fila.SKU
        if not isinstance(sku, str):
            continue
        linea_excel = fila.Index + header
        if sku in sku_lineas:
            sku_lineas[sku].append(linea_excel)
        else:
            sku_lineas[sku] = [linea_excel]
    return {sku: lineas for sku, lineas in sku_lineas.items() if len(lineas) > 1}


def previsualizar_lista(file, header: int, fecha, is_oferta: bool) -> dict:
    """Valida el Excel de lista de precios sin crear la tarea ni modificar precios.

    Pasos: detecta repetidos, datos inválidos y SKU ya programados (PostgreSQL),
    luego crea una tabla temporal DBISAM de solo lectura para verificar existencia en
    inventario y ofertas activas. La tabla temporal se elimina siempre al finalizar.

    Args:
        file: archivo Excel subido.
        header: fila del encabezado en el Excel (1-based).
        fecha: date de ejecución de la lista.
        is_oferta: True si la lista es de ofertas.

    Returns:
        dict con resumen de validación, o ValueError en caso de error crítico de lectura.
    """
    try:
        with pd.ExcelFile(file) as excel_file:
            df = pd.read_excel(excel_file, header=header - 1, dtype={'SKU': str})
        if df.empty:
            return ValueError("El archivo está vacío o no contiene datos válidos.")
        df = df.dropna(how='all')

        # 1. SKU repetidos dentro del archivo
        repetidos = detectar_skus_repetidos(df, header)

        # 2. Filas con datos inválidos (SKU no str o PRECIO no numérico)
        invalidos = []
        skus_invalidos: set = set()
        for fila in df.itertuples():
            if not isinstance(fila.SKU, str) or not isinstance(fila.PRECIO, (int, float)):
                invalidos.append({'sku': str(getattr(fila, 'SKU', 'N/A')), 'linea': fila.Index + header})
                if isinstance(fila.SKU, str):
                    skus_invalidos.add(fila.SKU)

        # SKUs únicos y válidos (primera aparición, sin repetidos ni inválidos) para DBISAM
        skus_repetidos = set(repetidos.keys())
        filas_validas = []
        skus_vistos: set = set()
        for fila in df.itertuples():
            sku = fila.SKU
            if not isinstance(sku, str) or sku in skus_invalidos or sku in skus_vistos:
                continue
            skus_vistos.add(sku)
            if sku not in skus_repetidos:
                filas_validas.append(fila)

        # 3. SKUs ya programados para la misma fecha (consulta PostgreSQL)
        skus_validos = [f.SKU for f in filas_validas]
        ya_programados = list(
            ProductsTasks.objects
            .filter(sku__in=skus_validos, fecha_ejecucion__date=fecha)
            .values_list('sku', flat=True)
        )
        ya_programados_set = set(ya_programados)

        # SKUs a insertar en la tabla temporal DBISAM (únicos, válidos, no ya programados)
        filas_a_insertar = [f for f in filas_validas if f.SKU not in ya_programados_set]

        con_oferta_activa: list = []
        sin_inventario: list = []
        pasaran: int = 0

        if filas_a_insertar:
            name_table = str(uuid4()).replace('-', '')
            dbisam_db = DBISAMDatabase()
            try:
                dbisam_db.create_table_tmp(name_table)
                for fila in filas_a_insertar:
                    dbisam_db.insert_data_tmp({
                        'Sku': fila.SKU,
                        'FechaInicio': fila.DESDE.strftime("%Y-%m-%d") if hasattr(fila, 'DESDE') else '1991-02-01',
                        'FechaFinal': fila.HASTA.strftime("%Y-%m-%d") if hasattr(fila, 'HASTA') else '1991-02-01',
                        'Precio': fila.PRECIO,
                        'PrecioAntes': fila.PRECIOANTES if hasattr(fila, 'PRECIOANTES') else fila.PRECIO - 2,
                        'Descripcion_Oferta': fila.DESCRIPCION if hasattr(fila, 'DESCRIPCION') else 'NO ES OFERTA',
                        'Tabla': name_table,
                    })
                # Completa DESCRIPCION/DEPARTAMENTO/CODBARRA/IVA (JOIN con SINVENTARIO)
                dbisam_db.update_table_tmp(name_table)
                # Productos sin inventario: DESCRIPCION queda NULL tras el INNER JOIN anterior
                sin_inventario = dbisam_db.get_skus_sin_inventario(name_table)
                # Marca ACTUALIZADO=TRUE a los que no tienen oferta activa
                fecha_str = fecha.strftime("%Y-%m-%d") if hasattr(fecha, 'strftime') else str(fecha)
                duplicados_dbisam = dbisam_db.update_tabla_tmp_productos_actualizados(
                    name_table, fecha_ejecucion=fecha_str
                )
                con_oferta_activa = [row[0] for row in duplicados_dbisam] if duplicados_dbisam else []
                pasaran = dbisam_db.get_productos_actualizados(name_table)
            finally:
                try:
                    dbisam_db.delete_table(name_table)
                except Exception:
                    pass

        return {
            'total': len(df),
            'pasaran': pasaran,
            'repetidos': {sku: lineas for sku, lineas in repetidos.items()},
            'invalidos': invalidos,
            'ya_programados': list(ya_programados),
            'con_oferta_activa': con_oferta_activa,
            'sin_inventario': sin_inventario,
        }

    except Exception as e:
        return ValueError(f"Error al previsualizar el archivo: {e}")


def read_excel_file(form, file, name_table, user, inmediato=False, is_oferta=False):
    try:
        with pd.ExcelFile(file) as excel_file:
            df = pd.read_excel(excel_file, header= form.cleaned_data['header'] - 1, dtype={'SKU': str})
            if df.empty:
                return ValueError("El archivo está vacio o no contiene datos válidos.")
            df = df.dropna(how='all')
            # Bloquear creación si hay SKU repetidos dentro del mismo archivo
            repetidos = detectar_skus_repetidos(df, form.cleaned_data['header'])
            if repetidos:
                msg = '; '.join(
                    f"'{sku}' en líneas {lineas}" for sku, lineas in repetidos.items()
                )
                return ValueError(
                    f"El archivo contiene SKU repetidos, corrija el Excel antes de continuar. "
                    f"SKU duplicados: {msg}"
                )
            dbisam_db = DBISAMDatabase()
            dbisam_db.create_table_tmp(name_table)
            fecha_ejecucion=datetime.combine(form.cleaned_data['date_time'], time.fromisoformat(settings.HORA_EJECUCION) )  #datetime.strptime(request.POST['date_time'], "%Y-%m-%d")
            list_products = []
            list_products_exists = []
            fecha_oferta = None
            #print('Fecha de ejecucion', fecha_ejecucion)
            for fila in df.itertuples():
                if not isinstance(fila.SKU, str) or not isinstance(fila.PRECIO, (int, float)):
                    return ValueError("El archivo debe contener datos válidos, los datos del producto %s en la linea %s" %(fila.SKU , fila.Index + form.cleaned_data['header']))
                
                if ProductsTasks.objects.filter(sku=fila.SKU, fecha_ejecucion=form.cleaned_data['date_time']).exists():
                        list_products_exists.append(fila.SKU)
                else:
                    fecha_oferta = fila.DESDE.strftime("%Y-%m-%d") if hasattr(fila, 'DESDE') else None
                    dbisam_db.insert_data_tmp({'Sku':fila.SKU, 
                                               'FechaInicio': fila.DESDE.strftime("%Y-%m-%d") if hasattr(fila, 'DESDE') else '1991-02-01',
                                               'FechaFinal': fila.HASTA.strftime("%Y-%m-%d") if hasattr(fila, 'HASTA') else '1991-02-01',
                                               'Precio':fila.PRECIO, 
                                               'PrecioAntes':fila.PRECIOANTES if hasattr(fila, 'PRECIOANTES') else fila.PRECIO - 2, 
                                               'Descripcion_Oferta': fila.DESCRIPCION if hasattr(fila, 'DESCRIPCION') else 'NO ES OFERTA',
                                               'Tabla':name_table})
                            
                    list_products.append(
                        dict(sku=fila.SKU, fecha_ejecucion=fecha_ejecucion,
                            is_oferta=is_oferta)
                                    )

                #if process:
                #    filas_validas.append(fila)
        
        if len(list_products) > 0:
            if not inmediato:
                if not is_oferta:
                    dbisam_db.update_table_tmp(name_table)
                
                    duplicados=dbisam_db.update_tabla_tmp_productos_actualizados(name_table, fecha_ejecucion=form.cleaned_data['date_time'])
                    row_count=dbisam_db.get_productos_actualizados(name_table)
                    list_products_exists.extend(duplicados)
                else:
                    dbisam_db.update_table_tmp(name_table)
                    duplicados=dbisam_db.update_tabla_tmp_productos_actualizados(name_table, fecha_ejecucion=fecha_oferta)
                    row_count=dbisam_db.get_productos_actualizados(name_table)
                    list_products_exists.extend(duplicados)
            else:
                
                dbisam_db.update_table_tmp(name_table)
                if not is_oferta:
                    row_count = dbisam_db.update_a2precios(name_table)
                    duplicados = dbisam_db.update_tabla_tmp_productos_actualizados(name_table)
                    list_products_exists.extend(duplicados)
                else:
                    duplicados=dbisam_db.update_tabla_tmp_productos_actualizados(name_table, fecha_ejecucion=fecha_oferta)
                    row_count = dbisam_db.insert_into_sinvoferta(name_table, list_products)  
                    list_products_exists.extend(duplicados)
            resaltar_sku_actualizados(file, list_products, header_row=form.cleaned_data['header'])
            return list_products, list_products_exists, row_count #filas_validas if process else len(df)
        else:
            dbisam_db.delete_table(name_table)
            return ValueError("Todos los productos encontrados en el archivo están asignados a una lista pendiente. Los siguientes SKU ya existen: %s" % ', '.join(list_products_exists))
    except Exception as e:  
        print(e)      
        return ValueError("Error al leer el archivo: {}".format(e))

   
def print_labels(row, request, rango_desde=None, rango_hasta=None):
    logo = request.POST.get('logo_etiquetas', '')
    formato_label = request.POST.get('formato_label', '')
    so_printers = request.POST.get('so_printers', '')
    if not logo or not formato_label or not so_printers:
        logger.warning(
            'print_labels: campos requeridos ausentes en POST '
            '(logo_etiquetas, formato_label, so_printers)'
        )
        return 0, []
    if rango_desde is not None and rango_hasta is not None:
        row = row[rango_desde - 1:rango_hasta]
    etiquetas_impresas = 0
    omitidos = []
    hPrinter = win32print.OpenPrinter(
                so_printers,
                {"DesiredAccess": win32print.PRINTER_ALL_ACCESS}
            )
    def generar_zpl(producto):
        match formato_label:
            case 'Ofertas':
                return PrinterLabel.label_zpl_ofertas({
                    'Logo': logo,
                    "SKU": f"{producto.SKU}",
                    "Descripcion": f"{producto.DESCRIPCION}",
                    "PrecioAntes": f"{producto.PRECIOANTES * ((producto.IVA / 100) + 1):.2f}",
                    "DescuentoPorcentaje": f"{round((producto.PRECIOANTES - producto.PRECIO) / producto.PRECIOANTES * 100, 2):.2f}",
                    "Precio": f"{producto.PRECIO * ((producto.IVA / 100) + 1):.2f}",
                    'Departamento': f"{producto.DEPARTAMENTO}",
                    "Desde": f"{producto.FECHA_INICIO.strftime('%d/%m/%Y')}",
                    "Hasta": f"{producto.FECHA_FINAL.strftime('%d/%m/%Y')}",
                })
            case 'Hablador':
                return PrinterLabel.label_zpl_hablador({
                    'Logo': logo,
                    "SKU": f"{producto.SKU}",
                    "Descripcion": f"{producto.DESCRIPCION}",
                    "PrecioAntes": f"{producto.PRECIOANTES * ((producto.IVA / 100) + 1):.2f}",
                    "CodBarra": f"{producto.CODBARRA}",
                    "Precio": f"{producto.PRECIO * ((producto.IVA / 100) + 1):.2f}",
                    'Departamento': f"{producto.DEPARTAMENTO}",
                })
            case _:
                return b""

    def chunked(iterable, size):
        """Divide cualquier iterable en lotes del tamaño indicado."""
        lote = []
        for item in iterable:
            lote.append(item)
            if len(lote) == size:
                yield lote
                lote = []
        if lote:  # Último lote si sobran elementos
            yield lote

    try:
        for numero_lote, lote in enumerate(chunked(row, 10), start=1):
            buffer = b""
            for producto in lote:
                motivo = validar_producto(producto, formato_label)
                if motivo:
                    omitidos.append((getattr(producto, 'SKU', 'SIN SKU'), motivo))
                    continue
                zpl = generar_zpl(producto)
                if zpl:
                    buffer += zpl
                    etiquetas_impresas += 1
            if not buffer:
                continue
            print(f"Enviando lote {numero_lote} — etiquetas ({etiquetas_impresas} total)")
            doc_abierto = False
            try:
                job_id = win32print.StartDocPrinter(hPrinter, 1, (f'ZPL_BATCH_{numero_lote}', None, 'RAW'))
                print(f"Lote {numero_lote}: job_id={job_id} (RAW)")
                doc_abierto = True
                win32print.StartPagePrinter(hPrinter)
                win32print.WritePrinter(hPrinter, buffer)
                win32print.EndPagePrinter(hPrinter)
                win32print.EndDocPrinter(hPrinter)
                doc_abierto = False
            except Exception as e:
                if doc_abierto:
                    try:
                        win32print.EndDocPrinter(hPrinter)
                    except Exception:
                        pass
                raise
            t.sleep(1.5)
    except Exception as e:
        print(f"Error de impresión: {e}")
        return 0, omitidos
    finally:
        try:
            win32print.ClosePrinter(hPrinter)
        except Exception:
            pass

    return etiquetas_impresas, omitidos


def validar_producto(producto, formato):
    if not isinstance(producto.PRECIO, (int, float)):
        return "PRECIO nulo o no numérico"
    if not isinstance(producto.PRECIOANTES, (int, float)):
        return "PRECIOANTES nulo o no numérico"
    if not isinstance(producto.IVA, (int, float)):
        return "IVA nulo o no numérico"
    if formato == 'Ofertas':
        if producto.PRECIOANTES == 0:
            return "PRECIOANTES es 0 (división por cero en el descuento)"
        if not hasattr(producto.FECHA_INICIO, 'strftime'):
            return "FECHA_INICIO nula o inválida"
        if not hasattr(producto.FECHA_FINAL, 'strftime'):
            return "FECHA_FINAL nula o inválida"
    return None


def resaltar_sku_actualizados(excel_path, lista_skus_actualizados, header_row, sku_col_name='SKU'):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    sku_col_idx = None
    for idx, cell in enumerate(ws[header_row]):
        if str(cell.value).strip().upper() == sku_col_name.strip().upper():
            sku_col_idx = idx
            break

    if sku_col_idx is None:
        print(f"No se encontró la columna '{sku_col_name}' en el encabezado.")
        wb.close()
        return

    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    green_font = Font(color='006100')
    lista_productos = [producto['sku'] for producto in lista_skus_actualizados]
    # Iterar sobre las filas debajo del encabezado
    for row in ws.iter_rows(min_row=header_row + 1):  # +2 porque openpyxl es 1-based y header_row es 0-based
        sku_cell = row[sku_col_idx]
        print(lista_skus_actualizados)
        if str(sku_cell.value) in lista_productos:
            sku_cell.fill = green_fill
            sku_cell.font = green_font

    wb.save(excel_path)
    wb.close()

def guardar_orden_lista(task, list_products: list[dict]) -> None:
    """Persiste el orden de las filas del Excel en ListaEtiquetaDetalle.

    Args:
        task: instancia de Tasks ya creada (con PK asignada).
        list_products: lista de dicts con clave 'sku' en el orden original del Excel
                       (retornada por read_excel_file como primer elemento de la tupla).
    """
    ListaEtiquetaDetalle.objects.bulk_create([
        ListaEtiquetaDetalle(task=task, orden=i, sku=fila['sku'])
        for i, fila in enumerate(list_products, start=1)
    ])


def generar_pdf():
       pass